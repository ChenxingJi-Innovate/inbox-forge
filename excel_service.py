"""Stateless Excel builder. No disk persistence.

Given a list of records, build an .xlsx in memory and return its bytes.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {
    "en": ["Date", "Customer", "Company", "Summary", "Follow-ups", "Status", "Created at"],
    "zh": ["日期", "客户名", "公司", "摘要", "待跟进", "状态", "创建时间"],
}
SHEET_TITLE = {"en": "Email Records", "zh": "邮件记录"}
COL_WIDTHS = [12, 16, 22, 60, 50, 12, 20]
HEADER_FILL_HEX = "E60023"  # pushpin-450


def _norm_lang(lang: str) -> str:
    return "zh" if str(lang or "").lower() == "zh" else "en"


def build_workbook(records: list[dict], lang: str = "en") -> bytes:
    """Construct an xlsx in memory and return its raw bytes."""
    lang = _norm_lang(lang)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_TITLE[lang]
    ws.append(HEADERS[lang])

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_HEX)
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for r in records or []:
        ws.append([
            r.get("date", ""),
            r.get("customer_name", ""),
            r.get("company", ""),
            r.get("summary", ""),
            r.get("action_items", ""),
            r.get("status", ""),
            r.get("created_at", ""),
        ])
        new_row = ws.max_row
        for col_idx in (4, 5):
            ws.cell(row=new_row, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
